"""Weekday vs Weekend analysis for a YouTube channel.

Generates a standalone HTML report comparing:
  (1) Publish-day effect: do videos published on weekdays accumulate more
      lifetime views than those published on weekends?
  (2) Viewing-day effect: across all videos, when do viewers actually watch?
      (Requires a YouTube Studio daily-views XLSX export.)

Data sources (in priority order):
  - YouTube Data API v3 via API key (set YOUTUBE_API_KEY + YOUTUBE_CHANNEL_ID).
    Public data only; no OAuth needed -> works for Brand Accounts.
  - Fallback: the most recent "Lifetime snapshot*.xlsx" in the parent folder.

Usage:
    python weekday-weekend/analyze.py
    open weekday-weekend/report.html
"""

from __future__ import annotations

import glob
import io
import json
import logging
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent

API_KEY = os.getenv("YOUTUBE_API_KEY", "").strip()
CHANNEL_ID = os.getenv("YOUTUBE_CHANNEL_ID", "").strip()
DAILY_XLSX = os.getenv("DAILY_XLSX", "").strip()  # optional Studio export
OUTPUT_HTML = HERE / "report.html"
TEMPLATE_HTML = HERE / "template.html"

WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


# --------------------------------------------------------------------------- #
# Data API v3 fetch (API key, no OAuth)
# --------------------------------------------------------------------------- #

def _api_get(endpoint: str, params: dict) -> dict:
    params = {**params, "key": API_KEY}
    url = f"https://www.googleapis.com/youtube/v3/{endpoint}"
    resp = requests.get(url, params=params, timeout=20)
    if resp.status_code != 200:
        log.error("API %s failed: %s %s", endpoint, resp.status_code, resp.text[:300])
        resp.raise_for_status()
    return resp.json()


def _get_uploads_playlist_id(channel_id: str) -> str | None:
    data = _api_get("channels", {"id": channel_id, "part": "contentDetails"})
    items = data.get("items", [])
    if not items:
        log.error("No channel found for id=%s", channel_id)
        return None
    return items[0]["contentDetails"]["relatedPlaylists"]["uploads"]


def _list_all_video_ids(uploads_playlist_id: str) -> list[str]:
    ids: list[str] = []
    page_token = None
    while True:
        params = {
            "playlistId": uploads_playlist_id,
            "part": "contentDetails",
            "maxResults": 50,
        }
        if page_token:
            params["pageToken"] = page_token
        data = _api_get("playlistItems", params)
        for item in data.get("items", []):
            vid = item["contentDetails"]["videoId"]
            ids.append(vid)
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    log.info("Found %d videos in uploads playlist", len(ids))
    return ids


_ISO_DUR = re.compile(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?")


def _iso_duration_to_seconds(iso: str) -> int:
    m = _ISO_DUR.match(iso or "")
    if not m:
        return 0
    h, mi, s = (int(x) if x else 0 for x in m.groups())
    return h * 3600 + mi * 60 + s


def _get_video_details(video_ids: list[str]) -> list[dict]:
    """Batch-fetch statistics + contentDetails + snippet for up to 50 ids at a time."""
    out: list[dict] = []
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i:i + 50]
        data = _api_get("videos", {
            "id": ",".join(batch),
            "part": "snippet,statistics,contentDetails",
            "maxResults": 50,
        })
        for item in data.get("items", []):
            snip = item.get("snippet", {})
            stats = item.get("statistics", {})
            dur = _iso_duration_to_seconds(item.get("contentDetails", {}).get("duration", ""))
            published = snip.get("publishedAt", "")
            pub_dt = None
            if published:
                try:
                    pub_dt = datetime.strptime(published[:10], "%Y-%m-%d")
                except ValueError:
                    pub_dt = None
            out.append({
                "id": item["id"],
                "title": snip.get("title", ""),
                "publish_date": pub_dt,
                "duration_sec": dur,
                "views": int(stats.get("viewCount", 0) or 0),
                "likes": int(stats.get("likeCount", 0) or 0),
                "comments": int(stats.get("commentCount", 0) or 0),
                "is_short": None,  # filled in by detect_shorts
            })
    return out


def _check_is_short(video_id: str) -> tuple[str, bool]:
    try:
        resp = requests.head(
            f"https://www.youtube.com/shorts/{video_id}",
            timeout=5,
            allow_redirects=False,
        )
        return video_id, resp.status_code == 200
    except Exception:
        return video_id, False


def _detect_shorts(video_ids: list[str]) -> dict[str, bool]:
    results: dict[str, bool] = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(_check_is_short, vid): vid for vid in video_ids}
        for f in as_completed(futures):
            vid, is_short = f.result()
            results[vid] = is_short
    return results


def fetch_from_api() -> list[dict] | None:
    if not (API_KEY and CHANNEL_ID):
        return None
    log.info("Fetching from YouTube Data API v3 for channel %s", CHANNEL_ID)
    uploads = _get_uploads_playlist_id(CHANNEL_ID)
    if not uploads:
        return None
    ids = _list_all_video_ids(uploads)
    videos = _get_video_details(ids)
    shorts = _detect_shorts(ids)
    for v in videos:
        v["is_short"] = shorts.get(v["id"], False)
    short_count = sum(1 for v in videos if v["is_short"])
    log.info("API: %d videos total (%d Shorts)", len(videos), short_count)
    return videos


# --------------------------------------------------------------------------- #
# Fallback: read the existing Lifetime snapshot xlsx
# --------------------------------------------------------------------------- #

def _parse_publish_date(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.strptime(s.strip(), "%b %d, %Y")
    except ValueError:
        return None


def fetch_from_xlsx() -> list[dict] | None:
    pattern = str(ROOT / "Lifetime snapshot*.xlsx")
    files = glob.glob(pattern)
    if not files:
        log.warning("No Lifetime snapshot xlsx found in %s", ROOT)
        return None
    path = max(files, key=os.path.getmtime)
    log.info("Reading xlsx: %s", path)
    wb = load_workbook(path)
    ws = wb["Table data"]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {
        "duration (seconds)": "duration_sec",
        "duration": "duration_sec",
        "likes": "likes",
        "comments added": "comments",
        "comments": "comments",
        "views": "views",
    }
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        k = header_map.get(str(h).strip().lower())
        if k:
            col_idx[k] = i

    def num(x) -> float:
        if x is None:
            return 0
        try:
            return float(x)
        except (ValueError, TypeError):
            return 0

    videos = []
    raw_ids = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        vid = row[0]
        if not vid or vid == "Total":
            continue
        raw_ids.append(vid)
        videos.append({
            "id": vid,
            "title": row[1] or "",
            "publish_date": _parse_publish_date(row[2]),
            "duration_sec": int(num(row[col_idx["duration_sec"]])) if "duration_sec" in col_idx else 0,
            "views": int(num(row[col_idx["views"]])) if "views" in col_idx else 0,
            "likes": int(num(row[col_idx["likes"]])) if "likes" in col_idx else 0,
            "comments": int(num(row[col_idx["comments"]])) if "comments" in col_idx else 0,
            "is_short": None,
        })
    wb.close()
    shorts = _detect_shorts(raw_ids)
    for v in videos:
        v["is_short"] = shorts.get(v["id"], False)
    log.info("xlsx: loaded %d videos", len(videos))
    return videos


# --------------------------------------------------------------------------- #
# Daily viewing xlsx (optional, for viewing-day analysis)
# --------------------------------------------------------------------------- #

def load_daily_views(path: str) -> list[tuple[date, int]] | None:
    """Parse a YouTube Studio 'daily views' XLSX export.

    Expected shape: a 'Table data' or first sheet with columns [Date, Views, ...].
    Returns list of (date, views) pairs.
    """
    if not path:
        return None
    p = Path(path)
    if not p.is_absolute():
        p = ROOT / path
    if not p.exists():
        log.warning("DAILY_XLSX not found: %s", p)
        return None
    log.info("Reading daily xlsx: %s", p)
    wb = load_workbook(p)
    ws = wb["Table data"] if "Table data" in wb.sheetnames else wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        date_col = headers.index("date")
    except ValueError:
        log.error("Daily xlsx missing 'Date' column. Headers: %s", headers)
        return None
    try:
        views_col = headers.index("views")
    except ValueError:
        log.error("Daily xlsx missing 'Views' column. Headers: %s", headers)
        return None

    out: list[tuple[date, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[date_col]
        v = row[views_col]
        if d is None or v is None:
            continue
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            try:
                d = datetime.strptime(d.strip(), "%Y-%m-%d").date()
            except ValueError:
                continue
        elif not isinstance(d, date):
            continue
        try:
            v = int(float(v))
        except (ValueError, TypeError):
            continue
        out.append((d, v))
    wb.close()
    log.info("Daily xlsx: %d day rows", len(out))
    return out


# --------------------------------------------------------------------------- #
# Analysis
# --------------------------------------------------------------------------- #

def build_publish_analysis(videos: list[dict]) -> dict:
    """Group by publish day-of-week and weekday/weekend."""
    dated = [v for v in videos if v.get("publish_date")]
    for v in dated:
        v["dow"] = v["publish_date"].weekday()  # 0 = Mon
        v["is_weekend"] = v["dow"] >= 5
        v["dow_name"] = WEEKDAY_NAMES[v["dow"]]

    def stats(group: list[dict]) -> dict:
        if not group:
            return {
                "count": 0, "total_views": 0, "avg_views": 0,
                "total_likes": 0, "avg_likes": 0,
                "total_comments": 0, "avg_comments": 0,
                "median_views": 0,
            }
        views = sorted(v["views"] for v in group)
        n = len(group)
        median = views[n // 2] if n % 2 else (views[n // 2 - 1] + views[n // 2]) / 2
        return {
            "count": n,
            "total_views": sum(v["views"] for v in group),
            "avg_views": round(sum(v["views"] for v in group) / n),
            "total_likes": sum(v["likes"] for v in group),
            "avg_likes": round(sum(v["likes"] for v in group) / n),
            "total_comments": sum(v["comments"] for v in group),
            "avg_comments": round(sum(v["comments"] for v in group) / n),
            "median_views": round(median),
        }

    by_dow = [stats([v for v in dated if v["dow"] == d]) for d in range(7)]
    weekday = stats([v for v in dated if not v["is_weekend"]])
    weekend = stats([v for v in dated if v["is_weekend"]])

    shorts = [v for v in dated if v["is_short"]]
    longform = [v for v in dated if not v["is_short"]]

    return {
        "total_videos": len(dated),
        "by_dow": by_dow,
        "weekday": weekday,
        "weekend": weekend,
        "weekday_shorts": stats([v for v in shorts if not v["is_weekend"]]),
        "weekend_shorts": stats([v for v in shorts if v["is_weekend"]]),
        "weekday_long": stats([v for v in longform if not v["is_weekend"]]),
        "weekend_long": stats([v for v in longform if v["is_weekend"]]),
        "by_dow_shorts": [stats([v for v in shorts if v["dow"] == d]) for d in range(7)],
        "by_dow_long": [stats([v for v in longform if v["dow"] == d]) for d in range(7)],
        "videos": [
            {
                "id": v["id"],
                "title": v["title"],
                "publish_date": v["publish_date"].strftime("%Y-%m-%d"),
                "dow_name": v["dow_name"],
                "is_weekend": v["is_weekend"],
                "is_short": v["is_short"],
                "duration_sec": v["duration_sec"],
                "views": v["views"],
                "likes": v["likes"],
                "comments": v["comments"],
                "thumbnail": f"https://i.ytimg.com/vi/{v['id']}/hqdefault.jpg",
                "url": f"https://www.youtube.com/watch?v={v['id']}",
            }
            for v in sorted(dated, key=lambda x: x["views"], reverse=True)
        ],
    }


def build_launch_analysis(
    videos: list[dict],
    daily: list[tuple[date, int]] | None,
) -> dict | None:
    """For each publish day of week, what does the channel's view curve look
    like in the 8 days starting from the publish day (day 0..7)?

    Because the daily XLSX is channel-wide (not per-video), this measures the
    aggregate view pattern around a publish event rather than a single video's
    performance. It's a reasonable proxy: if Thursday publishes consistently
    precede a workweek view surge, we'll see the curve peak on days 1-3.
    """
    if not daily:
        return None
    daily_map: dict[date, int] = {d: v for d, v in daily}
    d_start = min(daily_map)
    d_end = max(daily_map)

    buckets: dict[int, list[list[int]]] = {i: [] for i in range(7)}
    totals: dict[int, list[int]] = {i: [] for i in range(7)}
    sample_videos: dict[int, int] = {i: 0 for i in range(7)}

    for v in videos:
        pd_ = v.get("publish_date")
        if not pd_:
            continue
        pub = pd_.date() if hasattr(pd_, "date") else pd_
        if pub < d_start or pub > d_end:
            continue
        curve = []
        complete = True
        total = 0
        for offset in range(8):
            day = pub.fromordinal(pub.toordinal() + offset)
            if day not in daily_map:
                complete = False
                break
            val = daily_map[day]
            curve.append(val)
            total += val
        if not complete:
            continue
        dow = pub.weekday()
        buckets[dow].append(curve)
        totals[dow].append(total)
        sample_videos[dow] += 1

    def avg_curve(curves: list[list[int]]) -> list[int]:
        if not curves:
            return [0] * 8
        return [round(sum(c[i] for c in curves) / len(curves)) for i in range(8)]

    def median(vals: list[int]) -> int:
        if not vals:
            return 0
        vs = sorted(vals)
        n = len(vs)
        return vs[n // 2] if n % 2 else round((vs[n // 2 - 1] + vs[n // 2]) / 2)

    total_eligible = sum(sample_videos.values())
    if total_eligible == 0:
        return None

    return {
        "range_start": d_start.strftime("%Y-%m-%d"),
        "range_end": d_end.strftime("%Y-%m-%d"),
        "total_eligible": total_eligible,
        "by_dow": [
            {
                "dow": i,
                "dow_name": WEEKDAY_NAMES[i],
                "sample": sample_videos[i],
                "curve": avg_curve(buckets[i]),
                "avg_8day_total": round(sum(totals[i]) / len(totals[i])) if totals[i] else 0,
                "median_8day_total": median(totals[i]),
            }
            for i in range(7)
        ],
    }


def build_viewing_analysis(daily: list[tuple[date, int]] | None) -> dict | None:
    if not daily:
        return None
    by_dow = [[] for _ in range(7)]
    timeseries: list[dict] = []
    for d, v in daily:
        dow = d.weekday()
        by_dow[dow].append(v)
        timeseries.append({
            "date": d.strftime("%Y-%m-%d"),
            "dow": dow,
            "is_weekend": dow >= 5,
            "views": v,
        })

    def agg(vals: list[int]) -> dict:
        if not vals:
            return {"days": 0, "total": 0, "avg": 0}
        return {
            "days": len(vals),
            "total": sum(vals),
            "avg": round(sum(vals) / len(vals)),
        }

    by_dow_stats = [agg(vals) for vals in by_dow]
    weekday_vals = [v for d, v in daily if d.weekday() < 5]
    weekend_vals = [v for d, v in daily if d.weekday() >= 5]

    return {
        "range_start": min(d for d, _ in daily).strftime("%Y-%m-%d"),
        "range_end": max(d for d, _ in daily).strftime("%Y-%m-%d"),
        "total_days": len(daily),
        "by_dow": by_dow_stats,
        "weekday": agg(weekday_vals),
        "weekend": agg(weekend_vals),
        "timeseries": timeseries,
    }


# --------------------------------------------------------------------------- #
# Rendering
# --------------------------------------------------------------------------- #

def build_insights(publish: dict, viewing: dict | None, launch: dict | None) -> dict:
    """Compute headline stats for the hero banner and caption copy."""
    by_dow = publish["by_dow"]
    ranked = [
        (i, WEEKDAY_NAMES[i], by_dow[i])
        for i in range(7)
        if by_dow[i]["count"] >= 3  # ignore tiny-n days
    ]
    ranked.sort(key=lambda x: x[2]["avg_views"], reverse=True)
    best = ranked[0] if ranked else None
    worst = ranked[-1] if ranked else None

    weekday_stats = publish["weekday"]
    weekend_stats = publish["weekend"]

    view_ratio = None
    viewing_winner = None
    if viewing:
        wk_avg = viewing["weekday"]["avg"]
        we_avg = viewing["weekend"]["avg"]
        if we_avg:
            view_ratio = round(wk_avg / we_avg, 2)
        viewing_winner = "weekday" if wk_avg > we_avg else "weekend"

    recommendation = None
    if best:
        best_dow, best_name, best_stats = best
        is_weekend = best_dow >= 5
        if is_weekend:
            rec_detail = (
                f"{best_name} is the top publish day at {best_stats['avg_views']:,} avg "
                f"lifetime views/video — but weekend samples are small and your audience "
                f"watches on weekdays. Test, don't commit."
            )
        else:
            rec_detail = (
                f"Publishing on {best_name} delivers {best_stats['avg_views']:,} avg lifetime "
                f"views per video — roughly {round(best_stats['avg_views'] / max(weekday_stats['avg_views'], 1), 1)}× "
                f"your weekday norm. Videos land before the Tue–Fri viewing peak."
            )
        recommendation = {
            "headline": f"Publish on {best_name}",
            "detail": rec_detail,
            "best_day": best_name,
            "best_avg": best_stats["avg_views"],
            "best_sample": best_stats["count"],
        }

    # Cadence recommendation: pick top-N by avg_views among days with sample >= 3,
    # and flag days to avoid (bottom 2 overall, including small-sample days).
    cadence = None
    if ranked:
        top_ordered = [
            {"name": name, "avg": stats["avg_views"], "n": stats["count"], "small_sample": False}
            for _, name, stats in ranked
        ]
        all_ordered = sorted(
            [
                {
                    "name": WEEKDAY_NAMES[i],
                    "avg": by_dow[i]["avg_views"],
                    "n": by_dow[i]["count"],
                    "small_sample": by_dow[i]["count"] < 3,
                }
                for i in range(7)
                if by_dow[i]["count"] > 0
            ],
            key=lambda d: d["avg"],
            reverse=True,
        )
        avoid = sorted(all_ordered, key=lambda d: d["avg"])[:2]
        cadence = {
            "top_3": top_ordered[:3],
            "top_4": top_ordered[:4],
            "top_5": top_ordered[:5],
            "avoid": avoid,
        }

    return {
        "best": {"name": best[1], "avg": best[2]["avg_views"], "n": best[2]["count"]} if best else None,
        "worst": {"name": worst[1], "avg": worst[2]["avg_views"], "n": worst[2]["count"]} if worst else None,
        "weekday_avg_publish": weekday_stats["avg_views"],
        "weekend_avg_publish": weekend_stats["avg_views"],
        "view_ratio_wk_we": view_ratio,
        "viewing_winner": viewing_winner,
        "recommendation": recommendation,
        "cadence": cadence,
    }


def render(publish: dict, viewing: dict | None, launch: dict | None,
           insights: dict, source_label: str) -> str:
    template = TEMPLATE_HTML.read_text(encoding="utf-8")
    payload = {
        "publish": publish,
        "viewing": viewing,
        "launch": launch,
        "insights": insights,
        "source": source_label,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "weekday_names": WEEKDAY_NAMES,
    }
    return template.replace(
        "/*__DATA__*/{}",
        json.dumps(payload, default=str, separators=(",", ":")),
    )


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> int:
    videos = fetch_from_api()
    source = "YouTube Data API v3 (live)"
    if videos is None:
        log.info("No API key/channel id configured, falling back to xlsx")
        videos = fetch_from_xlsx()
        source = "Lifetime snapshot (xlsx)"
    if not videos:
        log.error("No data source available. Set YOUTUBE_API_KEY + YOUTUBE_CHANNEL_ID "
                  "or place a 'Lifetime snapshot*.xlsx' in the project root.")
        return 1

    publish = build_publish_analysis(videos)
    daily = load_daily_views(DAILY_XLSX) if DAILY_XLSX else None
    viewing = build_viewing_analysis(daily)
    launch = build_launch_analysis(videos, daily)
    insights = build_insights(publish, viewing, launch)
    html = render(publish, viewing, launch, insights, source)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    log.info("Wrote %s (%d videos, viewing data: %s, launch data: %s)",
             OUTPUT_HTML, publish["total_videos"],
             "yes" if viewing else "no",
             f"yes ({launch['total_eligible']} eligible videos)" if launch else "no")
    return 0


if __name__ == "__main__":
    sys.exit(main())

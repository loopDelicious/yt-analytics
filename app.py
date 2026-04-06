import io
import logging
import os
import glob
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from functools import wraps

import requests
from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", os.urandom(32).hex())

SITE_PASSWORD = os.getenv("SITE_PASSWORD", "changeme")
DATA_URL = os.getenv("DATA_URL", "")
CACHE_TTL = int(os.getenv("CACHE_TTL", "600"))

_cache = {"data": None, "expires": 0}


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def find_latest_snapshot():
    """Find the most recent 'Lifetime snapshot' xlsx in the working directory."""
    pattern = os.path.join(os.path.dirname(__file__) or ".", "Lifetime snapshot*.xlsx")
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def download_remote_snapshot():
    """Download xlsx from DATA_URL. Handles Google Drive confirmation pages."""
    if not DATA_URL:
        log.warning("No DATA_URL configured and no local xlsx found.")
        return None, None
    try:
        log.info("Downloading remote snapshot from DATA_URL...")
        sess = requests.Session()
        resp = sess.get(DATA_URL, timeout=30)
        resp.raise_for_status()

        # Google Drive may serve a virus-scan confirmation page for larger files.
        # Detect it and follow through with the confirm token.
        content_type = resp.headers.get("Content-Type", "")
        if "text/html" in content_type:
            log.info("Got HTML response, looking for Google Drive confirm token...")
            confirm_match = re.search(r'confirm=([0-9A-Za-z_-]+)', resp.text)
            uuid_match = re.search(r'uuid=([0-9A-Za-z_-]+)', resp.text)
            if confirm_match:
                params = {"confirm": confirm_match.group(1)}
                if uuid_match:
                    params["uuid"] = uuid_match.group(1)
                resp = sess.get(DATA_URL, params=params, timeout=30)
                resp.raise_for_status()
            else:
                # Try the id= param approach as fallback
                id_match = re.search(r'id=([0-9A-Za-z_-]+)', DATA_URL)
                if id_match:
                    download_url = f"https://drive.google.com/uc?export=download&confirm=t&id={id_match.group(1)}"
                    resp = sess.get(download_url, timeout=30)
                    resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "")
        if "text/html" in content_type:
            log.error("DATA_URL returned HTML instead of a file. Check your sharing settings.")
            log.error("Response preview: %s", resp.text[:500])
            return None, None

        disposition = resp.headers.get("Content-Disposition", "")
        match = re.search(r'filename\*?="?([^";\n]+)', disposition)
        filename = match.group(1).strip() if match else "Remote snapshot.xlsx"
        # Clean up URL-encoded filenames
        if filename.startswith("UTF-8''"):
            from urllib.parse import unquote
            filename = unquote(filename[7:])

        log.info("Downloaded %d bytes, filename: %s", len(resp.content), filename)
        return io.BytesIO(resp.content), filename
    except Exception as e:
        log.error("Failed to download from DATA_URL: %s", e)
        return None, None


def check_is_short(video_id):
    """Check if a video is a YouTube Short by probing the /shorts/ URL."""
    try:
        resp = requests.head(
            f"https://www.youtube.com/shorts/{video_id}",
            timeout=5,
            allow_redirects=False,
        )
        return video_id, resp.status_code == 200
    except Exception:
        return video_id, False


def detect_shorts(video_ids):
    """Batch-check which videos are Shorts using concurrent HEAD requests."""
    results = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        futures = {pool.submit(check_is_short, vid): vid for vid in video_ids}
        for future in as_completed(futures):
            vid, is_short = future.result()
            results[vid] = is_short
    short_count = sum(1 for v in results.values() if v)
    log.info("Shorts detection: %d/%d are Shorts", short_count, len(results))
    return results


def parse_publish_date(date_str):
    """Parse 'Mon DD, YYYY' format into a datetime object."""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str.strip(), "%b %d, %Y")
    except ValueError:
        return None


def load_video_data():
    """Load video data from YouTube API, local xlsx, or remote DATA_URL (in priority order)."""
    try:
        import youtube_api
        if youtube_api.is_available():
            log.info("YouTube API credentials found – fetching from API...")
            videos, snapshot_date_str, totals = youtube_api.fetch_channel_data()
            if videos:
                shorts_map = detect_shorts([v["id"] for v in videos])
                for v in videos:
                    v["is_short"] = shorts_map.get(v["id"], False)
                _cache["data"] = (videos, snapshot_date_str, totals)
                _cache["expires"] = time.time() + CACHE_TTL
                log.info("Cached API data for %d seconds", CACHE_TTL)
                return videos, snapshot_date_str, totals
            log.warning("YouTube API returned no videos, falling back to xlsx...")
    except ImportError:
        log.info("youtube_api module not available, skipping API source.")
    except Exception as e:
        log.error("YouTube API error: %s, falling back to xlsx...", e)

    filepath = find_latest_snapshot()

    if filepath:
        log.info("Loading local file: %s", filepath)
        filename = os.path.basename(filepath)
        wb = load_workbook(filepath)
    else:
        log.info("No local xlsx found, trying DATA_URL...")
        remote_data, filename = download_remote_snapshot()
        if remote_data is None:
            log.error("No data available. Set DATA_URL, configure YouTube API, or place an xlsx in the project root.")
            return [], "", None
        try:
            wb = load_workbook(remote_data)
        except Exception as e:
            log.error("Failed to parse downloaded xlsx: %s", e)
            return [], "", None

    match = re.search(r"Lifetime snapshot (.+)\.xlsx", filename or "")
    snapshot_date_str = match.group(1) if match else (filename or "Unknown")
    log.info("Snapshot date: %s", snapshot_date_str)
    ws = wb["Table data"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    raw_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        video_id = row[0]
        if not video_id or video_id == "Total":
            continue
        raw_rows.append(row)

    video_ids = [row[0] for row in raw_rows]
    shorts_map = detect_shorts(video_ids)

    videos = []
    for row in raw_rows:
        video_id = row[0]
        duration = row[3] or 0
        views = row[8] or 0
        subscribers = row[9] or 0
        videos.append({
            "id": video_id,
            "title": row[1] or "",
            "publish_time": row[2] or "",
            "publish_date": parse_publish_date(row[2]),
            "duration_sec": duration,
            "is_short": shorts_map.get(video_id, False),
            "likes": row[4] or 0,
            "comments": row[5] or 0,
            "avg_pct_viewed": row[6] or 0,
            "like_ratio": row[7] or 0,
            "views": views,
            "subscribers": subscribers,
            "subs_views_ratio": round(subscribers / views * 100, 2) if views else 0,
            "impressions": row[10] or 0,
            "ctr": row[11] or 0,
        })

    totals_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    totals = {
        "likes": totals_row[4] or 0,
        "comments": totals_row[5] or 0,
        "avg_pct_viewed": totals_row[6] or 0,
        "like_ratio": totals_row[7] or 0,
        "views": totals_row[8] or 0,
        "subscribers": totals_row[9] or 0,
        "impressions": totals_row[10] or 0,
        "ctr": totals_row[11] or 0,
    }

    wb.close()
    log.info("Loaded %d videos, total subs: %s", len(videos), totals.get("subscribers"))

    _cache["data"] = (videos, snapshot_date_str, totals)
    _cache["expires"] = time.time() + CACHE_TTL
    log.info("Cached data for %d seconds", CACHE_TTL)

    return videos, snapshot_date_str, totals


def get_video_data():
    """Return cached data if fresh, otherwise reload."""
    if _cache["data"] and time.time() < _cache["expires"]:
        return _cache["data"]
    return load_video_data()


def compute_averages(videos):
    """Compute per-video averages across all lifetime videos for color-coding."""
    keys = ["likes", "comments", "avg_pct_viewed", "like_ratio",
            "views", "subscribers", "impressions", "ctr", "subs_views_ratio"]
    if not videos:
        return {k: 0 for k in keys}
    n = len(videos)
    return {k: round(sum(v[k] for v in videos) / n, 2) for k in keys}


def filter_by_timespan(videos, days):
    """Filter videos published within the last N days."""
    if days is None:
        return videos
    cutoff = datetime.now() - timedelta(days=days)
    return [v for v in videos if v["publish_date"] and v["publish_date"] >= cutoff]


def filter_by_format(videos, fmt):
    """Filter by video format: 'all', 'shorts', or 'long'."""
    if fmt == "shorts":
        return [v for v in videos if v["is_short"]]
    if fmt == "long":
        return [v for v in videos if not v["is_short"]]
    return videos


def sort_videos(videos, tab, sort):
    """Sort videos descending by the primary metric for the given tab/sort."""
    if tab == "subscribers" and sort == "ratio":
        return sorted(videos, key=lambda v: v["subs_views_ratio"], reverse=True)
    if tab == "likability":
        return sorted(videos, key=lambda v: (v["like_ratio"], v["likes"]), reverse=True)
    sort_keys = {
        "subscribers": "subscribers",
        "engagement": "avg_pct_viewed",
    }
    key = sort_keys.get(tab, "subscribers")
    return sorted(videos, key=lambda v: v[key], reverse=True)


def format_number(n):
    """Format large numbers with commas."""
    if isinstance(n, float) and n == int(n):
        n = int(n)
    if isinstance(n, int):
        return f"{n:,}"
    return str(n)


def format_duration(seconds):
    """Format duration in seconds to MM:SS or HH:MM:SS."""
    if not seconds:
        return "0:00"
    seconds = int(seconds)
    if seconds >= 3600:
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return f"{h}:{m:02d}:{s:02d}"
    m = seconds // 60
    s = seconds % 60
    return f"{m}:{s:02d}"


app.jinja_env.filters["fmt_num"] = format_number
app.jinja_env.filters["fmt_dur"] = format_duration


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("password") == SITE_PASSWORD:
            session["authenticated"] = True
            return redirect(url_for("dashboard"))
        flash("Incorrect password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    tab = request.args.get("tab", "subscribers")
    timespan = request.args.get("timespan", "lifetime")
    sort = request.args.get("sort", "default")
    fmt = request.args.get("fmt", "all")

    if tab not in ("subscribers", "likability", "engagement"):
        tab = "subscribers"
    if sort not in ("default", "ratio"):
        sort = "default"
    if fmt not in ("all", "shorts", "long"):
        fmt = "all"

    days_map = {
        "lifetime": None,
        "7": 7,
        "28": 28,
        "90": 90,
        "365": 365,
    }
    days = days_map.get(timespan)

    empty_totals = {
        "likes": 0, "comments": 0, "avg_pct_viewed": 0, "like_ratio": 0,
        "views": 0, "subscribers": 0, "impressions": 0, "ctr": 0,
    }

    videos, snapshot_date, totals = get_video_data()
    if totals is None:
        totals = empty_totals
    averages = compute_averages(videos)
    filtered = filter_by_timespan(videos, days)
    filtered = filter_by_format(filtered, fmt)
    sorted_videos = sort_videos(filtered, tab, sort)

    if filtered:
        filtered_totals = {
            "likes": sum(v["likes"] for v in filtered),
            "comments": sum(v["comments"] for v in filtered),
            "avg_pct_viewed": round(sum(v["avg_pct_viewed"] for v in filtered) / len(filtered), 2),
            "like_ratio": round(sum(v["like_ratio"] for v in filtered) / len(filtered), 2),
            "views": sum(v["views"] for v in filtered),
            "subscribers": sum(v["subscribers"] for v in filtered),
            "impressions": sum(v["impressions"] for v in filtered),
            "ctr": round(sum(v["ctr"] for v in filtered) / len(filtered), 2),
        }
    else:
        filtered_totals = totals

    display_totals = filtered_totals if days is not None else totals

    return render_template(
        "dashboard.html",
        videos=sorted_videos,
        tab=tab,
        sort=sort,
        fmt=fmt,
        timespan=timespan,
        snapshot_date=snapshot_date,
        totals=display_totals,
        averages=averages,
        video_count=len(sorted_videos),
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)
